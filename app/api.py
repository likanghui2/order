import json
import os
import time
import uuid
import csv
from copy import deepcopy
from contextlib import asynccontextmanager
from datetime import date, datetime
from io import BytesIO, StringIO
from pathlib import Path
from typing import Any, Optional
from urllib.parse import unquote, urlparse

from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

from tools import nine_g_app_trace_token_producer, vj_web_session_server
from tools.vj_web_session_server import get_vj_session
from .runner import LocalRunner
from .source_registry import module_for_source, normalize_source, supported_sources
from .store import ACTIVE, DEFAULT_PRECHECK_RESOURCE_MISS_LIMIT, TaskStore


APP_DIR = Path(__file__).resolve().parents[1]
STATIC_DIR = APP_DIR / "static"
DB_PATH = Path(os.getenv("LOCAL_SHAM_DB", APP_DIR / "local_sham_booking.db"))
MAX_TABLE_IMPORT_ROWS = 1001
MAX_TABLE_IMPORT_COLUMNS = 50
TABLE_IMPORT_HEADERS = ["Source", "出发地", "目的地", "日期", "航班号", "舱位", "价格区间", "查询延迟", "预计延迟", "币种", "人数", "PNR有效期", "护照"]
TABLE_IMPORT_COLUMN_WIDTHS = [14, 12, 12, 14, 14, 10, 14, 12, 12, 10, 12, 14, 10]


class TaskPayload(BaseModel):
    task_id: Optional[str] = Field(default=None, alias="taskId")
    source: str
    task_type: str = Field(default="shamBooking", alias="taskType")
    task_data: dict[str, Any] = Field(..., alias="taskData")
    interval_seconds: Optional[int] = Field(default=None, ge=1, alias="intervalSeconds")
    max_runs: Optional[int] = Field(default=None, ge=1, alias="maxRuns")
    first_run_at: Optional[Any] = Field(default=None, alias="firstRunAt")
    passenger_range: Optional[str] = Field(default=None, alias="passengerRange")


class ImportPayload(BaseModel):
    tasks: Any
    replace_existing: bool = Field(default=True, alias="replaceExisting")


class TaskExportPayload(BaseModel):
    task_ids: list[str] = Field(default_factory=list, alias="taskIds")


class SourceProxyPayload(BaseModel):
    enabled: bool = False
    host: Optional[str] = None
    port: Optional[int] = Field(default=None, ge=1, le=65535)
    username: Optional[str] = None
    password: Optional[str] = None
    region: Optional[str] = None
    session_time: Optional[int] = Field(default=None, ge=1, alias="sessionTime")
    format: Optional[str] = None


class AppSettingsPayload(BaseModel):
    precheck_resource_miss_limit: Optional[int] = Field(default=None, ge=1, le=1000, alias="precheckResourceMissLimit")


store = TaskStore(DB_PATH)
runner = LocalRunner(
    store,
    concurrency=int(os.getenv("LOCAL_SHAM_CONCURRENCY", "0")),
    poll_interval=float(os.getenv("LOCAL_SHAM_POLL_INTERVAL", "0.5")),
)


@asynccontextmanager
async def lifespan(app: FastAPI):
    nine_g_app_trace_token_producer._start_producer()
    vj_web_session_server._start_warmer()
    runner.start()
    try:
        yield
    finally:
        runner.stop()
        vj_web_session_server._stop_warmer()
        nine_g_app_trace_token_producer._stop_producer()


app = FastAPI(title="Local Sham Booking", lifespan=lifespan)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True,
)


@app.get("/api/health")
def health():
    return {
        "status": "ok",
        "time": int(time.time()),
        "dbPath": str(DB_PATH),
        "runner": runner.stats(),
        "nineGTraceProducer": nine_g_app_trace_token_producer.producer_status(),
    }


@app.get("/api/sources")
def sources():
    return {"sources": supported_sources()}


@app.get("/api/settings")
def get_settings():
    return _settings_response(store.get_app_settings())


@app.put("/api/settings")
def save_settings(request: AppSettingsPayload):
    payload: dict[str, Any] = {}
    if request.precheck_resource_miss_limit is not None:
        payload["precheck_resource_miss_limit"] = request.precheck_resource_miss_limit
    return _settings_response(store.update_app_settings(payload))


@app.get("/api/source-proxies")
def list_source_proxies():
    return [_proxy_response(item) for item in store.list_source_proxy_configs(supported_sources())]


@app.put("/api/source-proxies/{source}")
def save_source_proxy(source: str, request: SourceProxyPayload):
    payload = _normalize_source_proxy_payload(source, request)
    return _proxy_response(store.upsert_source_proxy_config(payload))


@app.delete("/api/source-proxies/{source}")
def delete_source_proxy(source: str):
    normalized = normalize_source(source)
    return _proxy_response(store.delete_source_proxy_config(normalized))


@app.get("/api/vj-web-session")
def get_vj_web_session(
    dep_airport: str = Query(..., alias="depAirport"),
    arr_airport: str = Query(..., alias="arrAirport"),
):
    departure_place = _clean_optional(dep_airport).upper()
    arrival = _clean_optional(arr_airport).upper()
    if not departure_place or not arrival:
        raise HTTPException(status_code=400, detail="depAirport/arrAirport 不能为空")

    return get_vj_session(dep_airport=departure_place, arr_airport=arrival)


@app.get("/api/tasks")
def list_tasks():
    return store.list_tasks()


@app.get("/api/pnrs")
def list_pnrs(
    task_id: str = Query(default="", alias="taskId"),
    pnr: str = "",
    source: str = "",
    flight_number: str = Query(default="", alias="flightNumber"),
    cabin: str = "",
    currency_code: str = Query(default="", alias="currencyCode"),
    dep_airport: str = Query(default="", alias="depAirport"),
    arr_airport: str = Query(default="", alias="arrAirport"),
    dep_date: str = Query(default="", alias="depDate"),
    passenger_count: str = Query(default="", alias="passengerCount"),
    passengers: str = "",
    expired: str = "",
    limit: int = Query(default=200, ge=1, le=500),
    offset: int = Query(default=0, ge=0),
):
    filters = {
        "taskId": task_id,
        "pnr": pnr,
        "source": source,
        "flightNumber": flight_number,
        "cabin": cabin,
        "currencyCode": currency_code,
        "depAirport": dep_airport,
        "arrAirport": arr_airport,
        "depDate": dep_date,
        "passengerCount": passenger_count,
        "passengers": passengers,
        "expired": expired if expired in {"valid", "expired", "unknown"} else "",
    }
    rows = _build_pnr_rows(filters)
    page_rows = rows[offset : offset + limit]
    return {
        "rows": page_rows,
        "total": len(rows),
        "limit": limit,
        "offset": offset,
        "hasMore": offset + len(page_rows) < len(rows),
    }


@app.post("/api/tasks")
def create_task(request: TaskPayload):
    payload = _normalize_payload(request)
    if store.get_task(payload["task_id"]):
        raise HTTPException(status_code=409, detail="taskId 已存在")
    passenger_counts = _parse_passenger_range(request.passenger_range)
    if passenger_counts:
        return _create_task_tree(payload, request.passenger_range or "", passenger_counts)
    return store.create_task(payload)


@app.put("/api/tasks/{task_id}")
def update_task(task_id: str, request: TaskPayload):
    payload = _normalize_payload(request, fallback_task_id=task_id)
    return store.upsert_task(payload)


@app.post("/api/tasks/import")
def import_tasks(request: ImportPayload):
    task_items = _coerce_import_items(request.tasks)
    imported = []
    for item in task_items:
        payload = _normalize_payload(TaskPayload.model_validate(item))
        if request.replace_existing:
            imported.append(store.upsert_task(payload))
        else:
            imported.append(store.create_task(payload))
    return {"count": len(imported), "tasks": imported}


@app.post("/api/table-import/preview")
async def preview_table_import(file: UploadFile = File(...)):
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="表格文件为空")
    rows = _parse_table_file(file.filename or "", content)
    return {
        "filename": file.filename,
        "rows": rows,
    }


@app.get("/api/table-import/template")
def download_table_import_template():
    buffer = _build_table_import_template()
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="sham-booking-table-template.xlsx"'},
    )


@app.post("/api/tasks/export")
def export_tasks(request: TaskExportPayload):
    task_ids = _unique_task_ids(request.task_ids)
    tasks = _tasks_for_export(task_ids)
    rows = _export_rows_from_tasks(tasks)
    if not rows:
        raise HTTPException(status_code=400, detail="没有可导出的主任务")
    buffer = _build_table_import_workbook([TABLE_IMPORT_HEADERS, *rows], "当前任务导出")
    filename = f"sham-booking-tasks-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@app.get("/api/tasks/{task_id}")
def get_task(task_id: str):
    task = store.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    task["attempts"] = store.list_attempts(task_id)
    return task


@app.get("/api/tasks/{task_id}/attempts")
def list_attempts(task_id: str):
    if not store.get_task(task_id):
        raise HTTPException(status_code=404, detail="任务不存在")
    return store.list_attempts(task_id)


@app.post("/api/tasks/{task_id}/pause")
def pause_task(task_id: str):
    task = store.pause_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    return task


@app.post("/api/tasks/{task_id}/resume")
def resume_task(task_id: str):
    task = store.resume_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    return task


@app.post("/api/tasks/{task_id}/stop")
def stop_task(task_id: str):
    task = store.stop_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    return task


@app.post("/api/tasks/{task_id}/run-now")
def run_now(task_id: str):
    task = store.run_now(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")
    return task


@app.delete("/api/tasks/{task_id}")
def delete_task(task_id: str):
    if not store.delete_task(task_id):
        raise HTTPException(status_code=404, detail="任务不存在")
    return {"status": "ok", "taskId": task_id}


app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")


@app.get("/")
def index():
    return FileResponse(STATIC_DIR / "index.html")


def _build_pnr_rows(filters: dict[str, str]) -> list[dict[str, Any]]:
    seen: set[str] = set()
    rows: list[dict[str, Any]] = []
    now = time.time()
    for row in store.list_pnr_records():
        key = f"{row['taskId']}:{row['pnr']}"
        if key in seen:
            continue
        seen.add(key)
        if _matches_pnr_filters(row, filters, now):
            rows.append(row)
    rows.sort(key=lambda item: item.get("createdAt") or 0, reverse=True)
    return rows


def _matches_pnr_filters(row: dict[str, Any], filters: dict[str, str], now: float) -> bool:
    if filters["taskId"] and not _contains(row.get("taskId"), filters["taskId"]):
        return False
    if filters["pnr"] and not _contains(row.get("pnr"), filters["pnr"]):
        return False
    if filters["source"] and str(row.get("source") or "").upper() != filters["source"].upper():
        return False
    if filters["flightNumber"] and not _contains(row.get("flightNumber"), filters["flightNumber"]):
        return False
    if filters["cabin"] and not _contains(row.get("cabin"), filters["cabin"]):
        return False
    if filters["currencyCode"] and not _contains(row.get("currencyCode"), filters["currencyCode"]):
        return False
    if filters["depAirport"] and not _contains(row.get("depAirport"), filters["depAirport"]):
        return False
    if filters["arrAirport"] and not _contains(row.get("arrAirport"), filters["arrAirport"]):
        return False
    if filters["depDate"] and not _contains(f"{row.get('depDate')} {_format_dep_date(row.get('depDate'))}", filters["depDate"]):
        return False
    if filters["passengerCount"] and not _contains(row.get("passengerCount"), filters["passengerCount"]):
        return False
    if filters["passengers"] and not _contains(row.get("passengers"), filters["passengers"]):
        return False
    if filters["expired"] and _pnr_expiry_state(row.get("expiresAt"), now) != filters["expired"]:
        return False
    return True


def _pnr_expiry_state(expires_at: Any, now: float) -> str:
    expires_at_value = _safe_float(expires_at)
    if not expires_at_value:
        return "unknown"
    return "expired" if now >= expires_at_value else "valid"


def _format_dep_date(value: Any) -> str:
    text = str(value or "").strip()
    if len(text) == 8 and text.isdigit():
        return f"{text[:4]}-{text[4:6]}-{text[6:8]}"
    return text


def _contains(value: Any, query: str) -> bool:
    return str(query or "").lower() in str(value if value is not None else "").lower()


def _safe_float(value: Any) -> Optional[float]:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _normalize_payload(request: TaskPayload, fallback_task_id: Optional[str] = None) -> dict[str, Any]:
    source = normalize_source(request.source)
    if not source:
        raise HTTPException(status_code=400, detail="缺少必填字段：Source")
    task_id = request.task_id or fallback_task_id or _generated_task_id(request.task_data, source)
    _validate_required_task_fields(request.task_data)
    return {
        "task_id": task_id,
        "source": source,
        "task_type": request.task_type,
        "task_data": request.task_data,
        "interval_seconds": request.interval_seconds,
        "max_runs": request.max_runs,
        "first_run_at": request.first_run_at,
        "status": ACTIVE,
    }


def _validate_required_task_fields(task_data: dict[str, Any]) -> None:
    booking_config = task_data.get("bookingConfig") or {}
    ext = task_data.get("ext") or {}
    required = [
        (task_data.get("depAirport"), "出发地"),
        (task_data.get("arrAirport"), "目的地"),
        (task_data.get("depDate"), "日期"),
        (task_data.get("flightNumber"), "航班号"),
        (booking_config.get("currencyCode"), "币种"),
    ]
    missing = [label for value, label in required if not _clean_optional(value)]
    raw_pnr_valid_minutes = _first_present(ext, "pnrValidMinutes", "pnrValidityMinutes", "pnrValidMinute")
    if not _clean_optional(raw_pnr_valid_minutes):
        missing.append("PNR有效期")
    if missing:
        raise HTTPException(status_code=400, detail=f"缺少必填字段：{'、'.join(missing)}")
    pnr_valid_minutes = _safe_float(raw_pnr_valid_minutes)
    if pnr_valid_minutes is None or pnr_valid_minutes <= 0:
        raise HTTPException(status_code=400, detail="PNR有效期必须大于 0")


def _first_present(data: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        if key in data:
            return data.get(key)
    return None


def _create_task_tree(payload: dict[str, Any], passenger_range: str, passenger_counts: list[int]) -> dict[str, Any]:
    parent_payload = dict(payload)
    parent_task_id = parent_payload["task_id"]
    has_search_precheck = bool(module_for_source(parent_payload["source"], "search"))
    parent_payload["is_parent"] = True
    parent_payload["passenger_range"] = passenger_range or _format_passenger_range(passenger_counts)
    parent_payload["first_run_at"] = None
    parent_payload["status"] = ACTIVE
    parent = store.create_task(parent_payload)
    children = []

    try:
        if has_search_precheck:
            precheck_payload = dict(payload)
            precheck_payload["task_id"] = _precheck_task_id(parent_task_id)
            precheck_payload["task_type"] = "search"
            precheck_payload["parent_task_id"] = parent_task_id
            precheck_payload["child_index"] = 0
            precheck_payload["passenger_range"] = parent_payload["passenger_range"]
            precheck_payload["task_data"] = _search_task_data_from_sham(payload["task_data"])
            precheck_payload["max_runs"] = None
            children.append(store.create_task(precheck_payload))
        for index, passenger_count in enumerate(passenger_counts, start=1):
            child_payload = dict(payload)
            child_payload["task_id"] = _child_task_id(parent_task_id, passenger_count, index)
            child_payload["parent_task_id"] = parent_task_id
            child_payload["child_index"] = index
            child_payload["passenger_count"] = passenger_count
            child_payload["passenger_range"] = parent_payload["passenger_range"]
            child_payload["task_data"] = _task_data_for_passenger_count(payload["task_data"], passenger_count)
            if has_search_precheck:
                child_payload["next_run_at"] = None
            children.append(store.create_task(child_payload))
    except Exception:
        store.delete_task(parent_task_id)
        raise
    parent["children"] = children
    return parent


def _search_task_data_from_sham(task_data: dict[str, Any]) -> dict[str, Any]:
    booking_config = task_data.get("bookingConfig") or {}
    ext = task_data.get("ext") or {}
    return {
        "callbackData": task_data.get("callbackData") or {},
        "freightRateType": task_data.get("freightRateType") or "PT",
        "depAirport": task_data.get("depAirport"),
        "arrAirport": task_data.get("arrAirport"),
        "depDate": _format_dep_date(task_data.get("depDate")),
        "retDate": task_data.get("retDate") or "",
        "adultNumber": 1,
        "childNumber": 0,
        "currencyCode": booking_config.get("currencyCode"),
        "flightNumber": task_data.get("flightNumber"),
        "cabin": task_data.get("cabin"),
        "priceInterval": task_data.get("priceInterval"),
        "cabinLevel": ext.get("cabinLevel") or task_data.get("cabinLevel") or "Y",
        "privateCode": ext.get("privateCode") or task_data.get("privateCode") or [],
    }


def _task_data_for_passenger_count(task_data: dict[str, Any], passenger_count: int) -> dict[str, Any]:
    child_task_data = deepcopy(task_data)
    ext = dict(child_task_data.get("ext") or {})
    ext["passengerCount"] = passenger_count
    child_task_data["ext"] = ext
    return child_task_data


def _precheck_task_id(parent_task_id: str) -> str:
    base_id = f"{parent_task_id}-SEARCH"
    if not store.get_task(base_id):
        return base_id
    while True:
        candidate = f"{base_id}-{uuid.uuid4().hex[:4]}"
        if not store.get_task(candidate):
            return candidate


def _child_task_id(parent_task_id: str, passenger_count: int, index: int) -> str:
    base_id = f"{parent_task_id}-P{passenger_count}"
    if not store.get_task(base_id):
        return base_id
    while True:
        candidate = f"{base_id}-{index}-{uuid.uuid4().hex[:4]}"
        if not store.get_task(candidate):
            return candidate


def _parse_passenger_range(value: Optional[str]) -> list[int]:
    if value is None:
        return []
    text = value.strip()
    if not text:
        return []
    normalized = text.replace("~", "-").replace("至", "-").replace("，", ",")
    try:
        if "," in normalized:
            counts = [int(item.strip()) for item in normalized.split(",") if item.strip()]
        elif "-" in normalized:
            start_text, end_text = normalized.split("-", 1)
            start = int(start_text.strip())
            end = int(end_text.strip())
            if end < start:
                raise ValueError
            counts = list(range(start, end + 1))
        else:
            counts = [int(normalized)]
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="人数格式应为 1-1、1-4 或 1,2,3") from exc
    if not counts or any(count <= 0 for count in counts):
        raise HTTPException(status_code=400, detail="人数必须是大于 0 的整数")
    if len(set(counts)) != len(counts):
        raise HTTPException(status_code=400, detail="人数不能重复")
    if len(counts) > 100:
        raise HTTPException(status_code=400, detail="一次最多拆分 100 个子任务")
    return counts


def _format_passenger_range(counts: list[int]) -> str:
    if not counts:
        return ""
    if len(counts) == 1:
        return f"{counts[0]}-{counts[0]}"
    return f"{counts[0]}-{counts[-1]}"


def _coerce_import_items(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, str):
        value = json.loads(value)
    if isinstance(value, dict) and isinstance(value.get("tasks"), list):
        value = value["tasks"]
    if isinstance(value, dict):
        return [dict(task, taskId=task_id) for task_id, task in value.items()]
    if isinstance(value, list):
        return value
    raise HTTPException(status_code=400, detail="导入内容必须是数组、tasks 数组或 taskId 映射")


def _unique_task_ids(task_ids: list[str]) -> list[str]:
    result = []
    seen = set()
    for task_id in task_ids or []:
        text = str(task_id or "").strip()
        if not text or text in seen:
            continue
        seen.add(text)
        result.append(text)
    return result


def _tasks_for_export(task_ids: list[str]) -> list[dict[str, Any]]:
    if task_ids:
        tasks = []
        missing = []
        for task_id in task_ids:
            task = store.get_task(task_id)
            if not task:
                missing.append(task_id)
                continue
            tasks.append(task)
        if missing:
            raise HTTPException(status_code=404, detail=f"任务不存在：{missing[0]}")
        return tasks
    return store.list_tasks()


def _export_rows_from_tasks(tasks: list[dict[str, Any]]) -> list[list[Any]]:
    rows = []
    for task in tasks:
        if task.get("parent_task_id") or task.get("task_type") == "search":
            continue
        rows.append(_table_import_row_from_task(task))
    return rows


def _table_import_row_from_task(task: dict[str, Any]) -> list[Any]:
    task_data = task.get("task_data") or {}
    booking_config = task_data.get("bookingConfig") or {}
    ext = task_data.get("ext") or {}
    return [
        task.get("source") or "",
        task_data.get("depAirport") or "",
        task_data.get("arrAirport") or "",
        _format_dep_date(task_data.get("depDate")),
        task_data.get("flightNumber") or "",
        task_data.get("cabin") or "",
        task_data.get("priceInterval") or "",
        task.get("interval_seconds") or "",
        booking_config.get("bookRate") or "",
        booking_config.get("currencyCode") or "",
        _export_passenger_range(task, task_data),
        _first_present(ext, "pnrValidMinutes", "pnrValidityMinutes", "pnrValidMinute") or "",
        "是" if _export_use_passport(task_data) else "否",
    ]


def _export_passenger_range(task: dict[str, Any], task_data: dict[str, Any]) -> str:
    if task.get("passenger_range"):
        return str(task["passenger_range"])
    ext = task_data.get("ext") or {}
    passenger_count = task.get("passenger_count") or ext.get("passengerCount")
    if passenger_count:
        return f"{passenger_count}-{passenger_count}"
    return ""


def _export_use_passport(task_data: dict[str, Any]) -> bool:
    ext = task_data.get("ext") or {}
    if "usePassport" not in ext:
        return True
    value = ext.get("usePassport")
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() not in {"false", "0", "no", "否"}


def _build_table_import_template() -> BytesIO:
    rows = [
        TABLE_IMPORT_HEADERS,
        ["5JWEB", "SZX", "KUL", "2026-06-03", "AK127", "L", "", 30, 10, "MYR", "1-3", 120, "是"],
        ["VJWEB", "CAN", "SGN", "2026-06-07", "VJ3909", "H", "", 30, 10, "VND", "1-1", 60, "是"],
        ["8MWEB", "RGN", "CAN", "2026-08-11", "8M711", "", "200-350", 30, 10, "USD", "1-1", 60, "是"],
    ]
    return _build_table_import_workbook(rows, "押位任务导入")


def _build_table_import_workbook(rows: list[list[Any]], sheet_title: str) -> BytesIO:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="缺少 openpyxl 依赖，请更新 requirements-local.txt 后安装") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_title
    for row in rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="DBEAFE")
    header_font = Font(bold=True, color="1E293B")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for index, width in enumerate(TABLE_IMPORT_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _parse_table_file(filename: str, content: bytes) -> list[list[str]]:
    suffix = Path(filename or "").suffix.lower()
    if suffix in {".xlsx", ".xlsm"} or (not suffix and content.startswith(b"PK")):
        rows = _parse_xlsx_rows(content)
    elif suffix == ".xls":
        rows = _parse_xls_rows(content)
    elif suffix in {".csv", ".tsv", ".txt", ""}:
        rows = _parse_text_table_rows(filename, content)
    else:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx、.xlsm、.xls、.csv、.tsv、.txt 表格文件")
    if not rows:
        raise HTTPException(status_code=400, detail="表格文件没有可解析的数据")
    return rows


def _parse_xlsx_rows(content: bytes) -> list[list[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="缺少 openpyxl 依赖，请更新 requirements-local.txt 后安装") from exc

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Excel 文件解析失败，请确认文件格式") from exc
    sheet = workbook.active
    rows = []
    for row in sheet.iter_rows(max_row=MAX_TABLE_IMPORT_ROWS, max_col=MAX_TABLE_IMPORT_COLUMNS, values_only=True):
        rows.append([_table_cell_to_text(cell) for cell in row])
    workbook.close()
    return _trim_table_rows(rows)


def _parse_xls_rows(content: bytes) -> list[list[str]]:
    try:
        import xlrd
    except ImportError as exc:
        raise HTTPException(status_code=500, detail="缺少 xlrd 依赖，请更新 requirements-local.txt 后安装") from exc

    try:
        workbook = xlrd.open_workbook(file_contents=content)
        sheet = workbook.sheet_by_index(0)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="XLS 文件解析失败，请确认文件格式") from exc
    rows = []
    max_rows = min(sheet.nrows, MAX_TABLE_IMPORT_ROWS)
    max_cols = min(sheet.ncols, MAX_TABLE_IMPORT_COLUMNS)
    for row_index in range(max_rows):
        cells = []
        for column_index in range(max_cols):
            cell = sheet.cell(row_index, column_index)
            value = cell.value
            if cell.ctype == xlrd.XL_CELL_DATE:
                try:
                    parts = xlrd.xldate_as_tuple(value, workbook.datemode)
                    value = date(*parts[:3]) if parts[3:] == (0, 0, 0) else datetime(*parts)
                except Exception:
                    pass
            cells.append(_table_cell_to_text(value))
        rows.append(cells)
    return _trim_table_rows(rows)


def _parse_text_table_rows(filename: str, content: bytes) -> list[list[str]]:
    text = _decode_table_text(content)
    suffix = Path(filename or "").suffix.lower()
    sample = text[:4096]
    delimiter = "\t" if suffix == ".tsv" or sample.count("\t") >= sample.count(",") else None
    if delimiter is None:
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            reader = csv.reader(StringIO(text), dialect)
        except csv.Error:
            reader = csv.reader(StringIO(text))
    else:
        reader = csv.reader(StringIO(text), delimiter=delimiter)
    rows = []
    for index, row in enumerate(reader):
        if index >= MAX_TABLE_IMPORT_ROWS:
            break
        rows.append([_table_cell_to_text(cell) for cell in row[:MAX_TABLE_IMPORT_COLUMNS]])
    return _trim_table_rows(rows)


def _decode_table_text(content: bytes) -> str:
    for encoding in ("utf-8-sig", "gb18030", "big5", "latin-1"):
        try:
            return content.decode(encoding)
        except UnicodeDecodeError:
            continue
    return content.decode("utf-8", errors="replace")


def _table_cell_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, bool):
        return "是" if value else "否"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _trim_table_rows(rows: list[list[str]]) -> list[list[str]]:
    trimmed = []
    for row in rows:
        cells = [str(cell or "").strip() for cell in row[:MAX_TABLE_IMPORT_COLUMNS]]
        while cells and not cells[-1]:
            cells.pop()
        if cells or trimmed:
            trimmed.append(cells)
    while trimmed and not any(trimmed[-1]):
        trimmed.pop()
    return trimmed


def _generated_task_id(task_data: dict[str, Any], source: str) -> str:
    dep = _clean_part(task_data.get("depAirport"), "DEP")
    arr = _clean_part(task_data.get("arrAirport"), "ARR")
    flight = _clean_part(task_data.get("flightNumber"), "FLIGHT")
    dep_date = "".join(ch for ch in str(task_data.get("depDate") or "") if ch.isdigit())[:8]
    suffix = str(uuid.uuid4().int % 100000).zfill(5)
    return f"{source}-{dep}-{arr}-{flight}-{dep_date or 'DATE'}-{suffix}"


def _clean_part(value: Any, fallback: str) -> str:
    cleaned = "".join(ch for ch in str(value or "").upper() if ch.isascii() and ch.isalnum())
    return cleaned or fallback


def _normalize_source_proxy_payload(source: str, request: SourceProxyPayload) -> dict[str, Any]:
    normalized_source = normalize_source(source)
    data = request.model_dump(by_alias=False)
    host, port, username, password, format_value = _normalize_proxy_host(
        data.get("host"),
        data.get("port"),
        data.get("username"),
        data.get("password"),
        data.get("format"),
    )
    enabled = bool(data.get("enabled"))
    if enabled and (not host or port is None):
        raise HTTPException(status_code=400, detail="启用代理时必须填写代理 IP/Host 和端口")
    if enabled and (not username or not password):
        raise HTTPException(status_code=400, detail="启用代理时必须填写用户名和密码")
    if enabled and not format_value:
        raise HTTPException(status_code=400, detail="启用代理时必须填写 Format")
    return {
        "source": normalized_source,
        "enabled": enabled,
        "host": host,
        "port": port,
        "username": username,
        "password": password,
        "region": _clean_optional(data.get("region")),
        "session_time": data.get("session_time"),
        "format": format_value,
    }


def _normalize_proxy_host(
    host: Optional[str],
    port: Optional[int],
    username: Optional[str],
    password: Optional[str],
    format_value: Optional[str],
) -> tuple[str, Optional[int], str, str, str]:
    host_text = _clean_optional(host)
    username_text = _clean_optional(username)
    password_text = _clean_optional(password)
    format_text = _clean_optional(format_value)
    if not host_text:
        return "", port, username_text, password_text, format_text

    if "://" in host_text:
        parsed = urlparse(host_text)
        if parsed.hostname:
            host_text = parsed.hostname
        if parsed.port and port is None:
            port = parsed.port
        if parsed.username and not username_text:
            username_text = unquote(parsed.username)
        if parsed.password and not password_text:
            password_text = unquote(parsed.password)
        if not format_text:
            format_text = host or ""
    elif ":" in host_text and port is None:
        maybe_host, maybe_port = host_text.rsplit(":", 1)
        if maybe_port.isdigit():
            host_text = maybe_host
            port = int(maybe_port)
    return host_text, port, username_text, password_text, format_text


def _proxy_response(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "source": item["source"],
        "enabled": bool(item.get("enabled")),
        "host": item.get("host") or "",
        "port": item.get("port"),
        "username": item.get("username") or "",
        "password": item.get("password") or "",
        "region": item.get("region") or "",
        "sessionTime": item.get("session_time"),
        "format": item.get("format") or "",
        "updatedAt": item.get("updated_at"),
        "configured": bool(item.get("configured")),
    }


def _settings_response(item: dict[str, Any]) -> dict[str, Any]:
    return {
        "precheckResourceMissLimit": item.get("precheck_resource_miss_limit") or DEFAULT_PRECHECK_RESOURCE_MISS_LIMIT,
    }


def _clean_optional(value: Optional[Any]) -> str:
    return str(value).strip() if value is not None else ""
